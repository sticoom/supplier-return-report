"""openpyxl 报告输出：复刻手工模板版式。"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from engine.models import ReportData, SupplierResult

BOLD = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
# 结算清单版式字体（复刻枫悦 sheet：微软雅黑 16/12/11）
YH = "微软雅黑"
F_TITLE = Font(name=YH, size=16, bold=True)
F_NAME = Font(name=YH, size=12, bold=True)
F_HDR = Font(name=YH, size=11, bold=True)
F_DATA = Font(name=YH, size=12)
F_LABEL = Font(name=YH, size=12, bold=True)
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

BAD_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")   # Excel sheet 名非法字符
FIXED_SHEET_NAMES = ("供应商费用明细", "低于200清单", "季度累计", "买家备注复核清单",
                     "供应商批次合格率", "数据校验")
SUMMARY_HEADER = ["序号", "供应商", "应扣金额", "批次合格率", "是否签署质量协议(版本)",
                  "承担比例(考核系数)", "应承担金额", "备注"]
REASON_SUBS = ["DEFECTIVE \n(存在瑕疵)", "MISSING_PARTS\n (部分零件缺失)",
               "QUALITY_UNACCEPTABLE\n (质量未达到期望)"]
NOTE_TEXT = ("备注：数据来源于领星系统，亚马逊后台数据；如对数据有疑问，请及时联系德拉姆品质部。\n"
             "考核依据：依据《质量保证协议》中质量退货金额条款。")
SIGN_ROWS = [("采购部", "审核及优化意见简述："), ("品质部", "审核及优化意见简述："),
             ("供应链中心", "审核及优化意见简述："), ("总经理", "最终意见：")]


def short_name(supplier: str) -> str:
    s = supplier.strip()
    for suffix in ("有限公司", "有限责任公司", "公司"):
        if s.endswith(suffix) and len(s) > len(suffix):
            s = s[: -len(suffix)]
            break
    head = s[:5]                      # 行政区划前缀（东莞市/台州/石家庄市…）在头部
    for marker in ("市", "州"):        # 先找「市」（如苏州市），再退回「州」（如台州）
        idx = head.rfind(marker)
        if 1 <= idx < len(s) - 1:
            return s[idx + 1:]
    return s or supplier


def _sheet_title(base: str, used: set[str]) -> str:
    name = BAD_SHEET_CHARS.sub("-", base)[:31] or "供应商"
    candidate, n = name, 2
    while candidate in used:
        suffix = f"({n})"
        candidate = name[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def _style_header(ws, row, headers, start_col=1):
    for j, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row, column=j, value=h)
        c.font = BOLD
        c.alignment = CENTER
        c.border = THIN


def _quarter_label(month: str) -> str:
    y, m = month.split("-")           # 刻意内联：保持 report 不依赖 rules（任务解耦）
    return f"{y}-Q{(int(m) - 1) // 3 + 1}"


def _summary_sheet(wb: Workbook, data: ReportData, only_low200: bool) -> None:
    ws = wb.create_sheet("低于200清单" if only_low200 else "供应商费用明细")
    header = (SUMMARY_HEADER + ["所属季度"]) if only_low200 else SUMMARY_HEADER
    _style_header(ws, 1, header)
    pool = data.low200 if only_low200 else (data.suppliers + data.low200)
    ordered = sorted(pool, key=lambda s: (-s.deduction, s.supplier))
    for i, s in enumerate(ordered, start=2):
        note = "本月无验货" if s.pass_rate is None else ""
        vals = [i - 1, s.supplier, s.deduction,
                s.pass_rate, s.agreement, s.coefficient, s.undertaken, note]
        if only_low200:
            vals.append(_quarter_label(data.report_month))
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN
        ws.cell(row=i, column=3).number_format = "0.00"
        if s.pass_rate is not None:
            ws.cell(row=i, column=4).number_format = "0.0%"
        ws.cell(row=i, column=6).number_format = "0.##"
        ws.cell(row=i, column=7).number_format = "0.00"
    ws.column_dimensions["B"].width = 34


def _supplier_sheet(wb: Workbook, used: set[str], data: ReportData,
                    s: SupplierResult) -> None:
    """供应商结算清单——版式逐项复刻 2026-7 人工结果「枫悦」sheet。"""
    ws = wb.create_sheet(_sheet_title(short_name(s.supplier), used))
    y, m = data.report_month.split("-")
    def _cell(row, col, value=None, font=None, align=None, numfmt=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font or F_DATA
        if align:
            c.alignment = align
        if numfmt:
            c.number_format = numfmt
        return c

    def _border_rows(r1, r2):
        for rr in range(r1, r2 + 1):
            for cc in range(1, 12):
                ws.cell(row=rr, column=cc).border = THIN

    # 列宽 / 行高（参考枫悦 sheet 实测值）
    for col, w in {"A": 13.7, "B": 23.7, "C": 15.6, "D": 12.1, "E": 15.6, "F": 18.8,
                   "G": 28.5, "H": 12.8, "I": 15.2, "J": 14.0, "K": 17.6}.items():
        ws.column_dimensions[col].width = w

    # 标题 + 供应商名称（各占一整行合并）
    ws.merge_cells("A1:K1")
    _cell(1, 1, f" {y} 年 {int(m)} 月供应商质量退货金额汇总表", F_TITLE, CENTER)
    ws.merge_cells("A2:K2")
    _cell(2, 1, f"供应商名称：{s.supplier}", F_NAME, LEFT)
    ws.row_dimensions[1].height = 45
    ws.row_dimensions[2].height = 45

    # 两行表头
    main = ["序号", "SKU名称", "销量", "退货量", "质量退货量（按亚马逊平台退货描述）",
            None, None, "质量退货率", "产品采购单价\n（元）", "质量退货金额\n（元）", "备注"]
    for j, h in enumerate(main, start=1):
        if h is not None:
            _cell(3, j, h, F_HDR, CENTER)
    for j, h in enumerate(REASON_SUBS, start=5):
        _cell(4, j, h, F_HDR, CENTER)
    for rng in ("A3:A4", "B3:B4", "C3:C4", "D3:D4", "E3:G3",
                "H3:H4", "I3:I4", "J3:J4", "K3:K4"):
        ws.merge_cells(rng)
    ws.row_dimensions[3].height = 39
    ws.row_dimensions[4].height = 39

    # 数据行：全部居中（含备注列）；质量退货率写成 Excel 公式
    r = 5
    for i, ln in enumerate(s.skus, start=1):        # skus 已只含质量退货量>0 的行
        vals = [i, ln.sku, ln.sales_qty, ln.return_qty, ln.qty_defective,
                ln.qty_missing_parts, ln.qty_quality_unacceptable,
                f"=SUM(G{r}+F{r}+E{r})/C{r}", ln.unit_price, ln.amount, ln.note or None]
        def _fmt(j):
            return ("0" if j in (3, 4, 5, 6, 7)
                    else "0.00%" if j == 8
                    else "0.00" if j in (9, 10)
                    else "0_ " if j == 11 else None)
        for j, v in enumerate(vals, start=1):
            _cell(r, j, v, F_DATA, CENTER, _fmt(j))
        r += 1
    last = r - 1 if r > 5 else 5

    # 汇总区：统计金额 / 是否签署 / 合格率 / 考核系数 / 考核金额
    r_sum = r
    ws.merge_cells(f"A{r_sum}:I{r_sum}")
    _cell(r_sum, 1, "统计金额：", F_LABEL, RIGHT)
    _cell(r_sum, 10, f"=SUM(J5:J{last})", F_DATA, CENTER, "0.00_ ")
    r_coef = r_sum + 1
    ws.merge_cells(f"A{r_coef}:D{r_coef}")
    _cell(r_coef, 1, "是否签署最新质量协议", F_LABEL, CENTER)
    ws.merge_cells(f"E{r_coef}:F{r_coef}")
    # 与手工版式一致：签了写「是」，没签/未匹配写「否」（版本信息在汇总表「是否签署质量协议(版本)」列）
    _cell(r_coef, 5, "否" if s.agreement in ("否", "未匹配协议") else "是",
          F_DATA, CENTER, "0%")
    ws.merge_cells(f"G{r_coef}:I{r_coef}")
    _cell(r_coef, 7, "考核系数：", F_LABEL, RIGHT)
    _cell(r_coef, 10, s.coefficient, F_DATA, CENTER)
    r_pass = r_coef + 1
    ws.merge_cells(f"A{r_pass}:D{r_pass}")
    _cell(r_pass, 1, "当月检验合格率：", F_LABEL, CENTER)
    ws.merge_cells(f"E{r_pass}:F{r_pass}")
    _cell(r_pass, 5, "/" if s.pass_rate is None else s.pass_rate,
          F_DATA, CENTER, "0.0%")
    ws.merge_cells(f"G{r_pass}:I{r_pass}")
    _cell(r_pass, 7, "考核金额：", F_LABEL, RIGHT)
    _cell(r_pass, 10, f"=J{r_coef}*J{r_sum}", F_DATA, CENTER, "0.00_ ")

    # 备注行（一格两行）+ 签字栏
    r_note = r_pass + 1
    ws.merge_cells(f"A{r_note}:K{r_note}")
    _cell(r_note, 1, NOTE_TEXT, F_LABEL,
          Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[r_note].height = 46
    for k, (dept, label) in enumerate(SIGN_ROWS):
        rr = r_note + 1 + k
        _cell(rr, 1, dept, F_DATA, CENTER)
        _cell(rr, 2, label, F_DATA, CENTER)
        ws.merge_cells(f"C{rr}:K{rr}")
        _cell(rr, 3, " " * 42, F_DATA, CENTER)
        ws.row_dimensions[rr].height = 46
    # 边框最后统一写：合并之后给每个格子（含 MergedCell）赋边框，
    # openpyxl 会把样式写进 XML（Excel 正常显示；openpyxl 读回时不显示是读取端特性）
    _border_rows(1, r_note + len(SIGN_ROWS))


def _quarterly_sheet(wb: Workbook, data: ReportData) -> None:
    q = wb.create_sheet("季度累计")
    _style_header(q, 1, ["月份", "供应商", "应扣金额", "应承担金额", "备注"])
    for i, row in enumerate(data.quarterly, start=2):
        vals = [row.month, row.supplier, row.deduction, row.undertaken,
                "季度小计" if row.is_subtotal else ""]
        for j, v in enumerate(vals, start=1):
            c = q.cell(row=i, column=j, value=v)
            if row.is_subtotal:
                c.font = BOLD
        q.cell(row=i, column=3).number_format = "0.00"
        q.cell(row=i, column=4).number_format = "0.00"


def _review_sheet(wb: Workbook, data: ReportData) -> None:
    rv = wb.create_sheet("买家备注复核清单")
    _style_header(rv, 1, ["订单号", "SKU", "退货原因", "买家备注", "退货时间", "供应商"])
    for i, row in enumerate(data.review, start=2):
        for j, v in enumerate([row.order_id, row.sku, row.reason_raw,
                               row.buyer_comment, row.return_time, row.supplier], start=1):
            rv.cell(row=i, column=j, value=v)


def _batch_matrix_sheet(wb: Workbook, data: ReportData) -> None:
    """第 4 节「供应商批次合格率」：供应商×月份矩阵（总批数/不合格批数/批次合格率），
    供应商用全名，品质部可从本表复制粘贴更新 wiki《2026年验货数据报表》的「供应商」sheet。"""
    bm = wb.create_sheet("供应商批次合格率")
    months = sorted({mo for cells in data.batch_matrix.values() for mo in cells})
    _style_header(bm, 1, ["供应商"])
    for j, mo in enumerate(months):
        c0 = 2 + j * 3                          # 每月 3 子列：总批数/不合格批数/批次合格率
        _style_header(bm, 1, [mo], start_col=c0)
        _style_header(bm, 2, ["总批数", "不合格批数", "批次合格率"], start_col=c0)
        bm.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + 2)
    bm.merge_cells("A1:A2")
    for i, sup in enumerate(sorted(data.batch_matrix), start=3):
        c = bm.cell(row=i, column=1, value=sup)
        c.border = THIN
        for j, mo in enumerate(months):
            cell = data.batch_matrix[sup].get(mo)
            if cell is None:                     # 该月无验货 → 留空
                continue
            total, failed = cell
            for k, v in enumerate((total, failed, (total - failed) / total)):
                cc = bm.cell(row=i, column=2 + j * 3 + k, value=v)
                cc.border = THIN
            bm.cell(row=i, column=2 + j * 3 + 2).number_format = "0.0%"
    if months:
        bm.cell(row=len(data.batch_matrix) + 4, column=1,
                value="备注：从验货原始数据计算，批次合格率＝合格批数÷总批数（供应商全名）。")
    bm.column_dimensions["A"].width = 34


def _validation_sheet(wb: Workbook, data: ReportData) -> None:
    vd = wb.create_sheet("数据校验")
    _style_header(vd, 1, ["类型", "明细"])
    for i, item in enumerate(data.validation, start=2):
        vd.cell(row=i, column=1, value=item.kind)
        vd.cell(row=i, column=2, value=item.detail)


def write_report(out_dir: str, data: ReportData) -> tuple[str, str]:
    """生成两个工作簿（用户拍板 2026-08-20：低于200 与正常分开）：

    ① 主工作簿：供应商费用明细 + ≥200 供应商结算清单 + 季度累计/复核/批次合格率/数据校验
    ② 低于200工作簿：低于200清单 + <200 供应商结算清单
    返回 (主工作簿路径, 低于200工作簿路径)。
    """
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    y, m = data.report_month.split("-")
    path = out / f"{y}年{int(m)}月供应商质量退货金额汇总表.xlsx"
    low_path = out / f"{y}年{int(m)}月供应商质量退货金额汇总表（低于200）.xlsx"

    wb = Workbook()
    wb.remove(wb.active)
    used = set(FIXED_SHEET_NAMES)
    _summary_sheet(wb, data, only_low200=False)
    for s in sorted(data.suppliers, key=lambda x: x.supplier):
        _supplier_sheet(wb, used, data, s)
    _quarterly_sheet(wb, data)
    _review_sheet(wb, data)
    _batch_matrix_sheet(wb, data)
    _validation_sheet(wb, data)
    wb.save(path)

    wb_low = Workbook()
    wb_low.remove(wb_low.active)
    used_low = set(FIXED_SHEET_NAMES)
    _summary_sheet(wb_low, data, only_low200=True)
    for s in sorted(data.low200, key=lambda x: x.supplier):
        _supplier_sheet(wb_low, used_low, data, s)
    wb_low.save(low_path)
    return str(path), str(low_path)


def write_supplier_workbook(path, data: ReportData, s: SupplierResult) -> str:
    """单供应商工作簿（PDF 转换的中间产物）。

    与 xlsx 报告的差异（用户拍板 2026-08-20）：
    ① 备注列的「按交货比例分摊」在 PDF 里不出现；
    ② 页面设为横向 + 宽度适配一页（竖版 A4 会把右侧列截掉）。
    """
    from openpyxl.worksheet.properties import PageSetupProperties
    wb = Workbook()
    wb.remove(wb.active)
    _supplier_sheet(wb, set(), data, s)
    ws = wb.active
    for row in ws.iter_rows(min_row=5):        # ① 抹掉分摊备注
        if row[10].value == "按交货比例分摊":
            row[10].value = None
    ws.page_setup.orientation = "landscape"    # ② 横向 + 缩放适配页宽
    ws.page_setup.paperSize = 9                # A4
    # LibreOffice 对 fitToWidth 支持不稳定 → 用固定缩放：
    # 全表 11 列约 14.7in 宽，A4 横向可用约 10.9in → 72% 恰好放下
    ws.page_setup.scale = 72
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    wb.save(path)
    return str(path)
