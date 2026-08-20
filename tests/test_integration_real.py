"""T7 端到端集成测试。

分两层：

1. 合成全量 E2E（默认跑）：用 conftest 的合成 7 月文件（openpyxl 造数，列名/文件名
   与真实导出一致）走真实 CLI（`python -m engine.pipeline` 子进程）全量跑出报告，
   再做「人工抽查 3 个供应商」的自动化等价：
     抽查1（甲）：结算清单「当月检验合格率」与验货 xlsx 原始单元格独立重算一致；
     抽查2（甲/乙/丙）：每条有价 SKU 行 件数×单价=金额，件数与 FBA/FBM 订单独立重算一致；
     抽查3（乙/丙）：交货拆分行 note=「按交货比例分摊」、小数件数、两行件数之和守恒。

2. 真实数据 E2E（marker=integration）：读 D:/Downloads 7 月真实文件 + 本地参考数据，
   合格率与验货文件独立重算交叉校验。默认被 pyproject addopts 的
   ``-m "not integration"`` 排除，须显式 ``pytest -m integration`` 且真实文件在本机
   才执行（避免普通单测运行触碰 123MB 验货报表）。
"""
from __future__ import annotations

import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

from conftest import (AGREEMENT_NAME, BING, DLM_NAME, FBA_NAME, FBM_NAME,
                      INBOUND_NAME, INSPECTION_NAME, JIA, MONTH, YI,
                      make_monthly_files, make_reference_files)
from engine import loaders, pipeline
from engine.models import ReferenceData

REPO_ROOT = Path(__file__).resolve().parent.parent
REPORT_NAME = "2026年7月供应商质量退货金额汇总表.xlsx"
LOW_REPORT_NAME = "2026年7月供应商质量退货金额汇总表（低于200）.xlsx"
QUALITY_CODES = ("DEFECTIVE", "MISSING_PARTS", "QUALITY_UNACCEPTABLE")

# 真实样例文件（只读参考，不入库；主 Agent 集成验证用）
REAL = {
    "fba": Path(r"D:/Downloads/退货(FBA)订单导出-947873100663549952.xlsx"),
    "fbm": Path(r"D:/Downloads/退货(FBM)订单导出-947873572875407360.xlsx"),
    "dlm": Path(r"D:/Downloads/DLM退货统计SKU导出-2026-08-17.xlsx"),
    "inbound": Path(r"C:/Users/13676/Desktop/飞书下载/采购入库单_202601-07.xlsx"),
    "inspection": Path(r"D:/Warp/供应商退货统计/data/2026年验货数据报表.xlsx"),
    "agreements": Path(r"D:/Warp/供应商退货统计/data/供应商协议签订记录.csv"),
}


# ---------- 报告读取 / 独立重算辅助（openpyxl 直读单元格，不经过 engine 规则） ----------

def _supplier_sheet(cli_run, supplier: str):
    """按 A2 标签定位供应商结算清单 sheet（先主工作簿，再低于200工作簿）。"""
    tag = f"供应商名称：{supplier}"
    for book in (cli_run.report, cli_run.low_report):
        if not Path(book).exists():
            continue
        for ws in openpyxl.load_workbook(book).worksheets:
            if ws["A2"].value == tag:
                return ws
    raise AssertionError(f"报告中找不到 {supplier} 的结算清单 sheet")


def _sku_rows(ws):
    """读结算清单数据行（第 5 行起，到「统计金额：」为止）→ dict 列表。"""
    rows = []
    r = 5
    while ws.cell(row=r, column=1).value != "统计金额：" and r <= ws.max_row:
        if ws.cell(row=r, column=2).value:            # B 列 = SKU名称
            rows.append({
                "sku": ws.cell(row=r, column=2).value,
                "qty_def": ws.cell(row=r, column=5).value or 0,     # E
                "qty_mp": ws.cell(row=r, column=6).value or 0,      # F
                "qty_qu": ws.cell(row=r, column=7).value or 0,      # G
                "price": ws.cell(row=r, column=9).value,            # I（缺价为 None）
                "amount": ws.cell(row=r, column=10).value,          # J（缺价为 None）
                "note": ws.cell(row=r, column=11).value or "",      # K
            })
        r += 1
    return rows


def _labeled_value(ws, label: str, col: int):
    """在 A 列找以 label 开头的行，返回该行 col 列的值（如「当月检验合格率」→E）。"""
    for r in range(5, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").startswith(label):
            return ws.cell(row=r, column=col).value
    raise AssertionError(f"sheet 中找不到标签行：{label}")


def _recount_pass_rate(inspection_path, supplier: str, month: str,
                       sheet: str = "26年验货原始数据"):
    """独立重算批次合格率：openpyxl 直读验货 xlsx 原始单元格数 合格/总批数。"""
    wb = openpyxl.load_workbook(inspection_path, read_only=True)
    try:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        i_sup, i_mon, i_res = (header.index("供应商"), header.index("月份"),
                               header.index("质检结果"))
        total = ok = 0
        for r in rows:
            sup = str(r[i_sup] or "").strip()
            mon = r[i_mon]
            mon = mon.strftime("%Y-%m") if isinstance(mon, datetime) else str(mon or "").strip()
            if sup == supplier.strip() and mon == month:
                total += 1
                if str(r[i_res] or "").strip() == "合格":
                    ok += 1
        return ok / total if total else None
    finally:
        wb.close()


def _recount_quality_qty(fba_path, fbm_path, sku: str) -> dict[str, float]:
    """独立重算某 SKU 三原因质量退货件数：openpyxl 直读 FBA/FBM 原始订单行，
    按「退货原因前缀」人工筛选求和（FBA 无前缀、FBM 带 CR- 前缀）。"""
    counts = dict.fromkeys(QUALITY_CODES, 0)

    def _tally(path, sku_col, prefixes):
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            ws = wb.worksheets[0]
            rows = ws.iter_rows(values_only=True)
            header = [str(c).strip() if c is not None else "" for c in next(rows)]
            i_sku, i_qty, i_reason = (header.index(sku_col), header.index("退货数量"),
                                      header.index("退货原因"))
            for r in rows:
                if str(r[i_sku] or "").strip() != sku:
                    continue
                reason = str(r[i_reason] or "").strip()
                for code, prefix in prefixes.items():
                    if reason.startswith(prefix):
                        counts[code] += float(r[i_qty] or 0)
                        break
        finally:
            wb.close()

    _tally(fba_path, "sku", {c: c for c in QUALITY_CODES})
    _tally(fbm_path, "SKU", {c: f"CR-{c}" for c in QUALITY_CODES})
    return counts


# ---------- 第 1 层：合成数据 CLI 全量 + 人工抽查 3 供应商（默认跑） ----------

@pytest.fixture(scope="module")
def cli_run(tmp_path_factory):
    """真实 CLI 子进程全量跑合成 7 月数据：--db/--out 全部隔离在 tmp，不碰项目 data/。"""
    tmp = tmp_path_factory.mktemp("cli_e2e")
    monthly = make_monthly_files(tmp / "files")
    refs = make_reference_files(tmp / "files")
    cmd = [sys.executable, "-m", "engine.pipeline",
           "--month", MONTH,
           "--fba", str(monthly[FBA_NAME]), "--fbm", str(monthly[FBM_NAME]),
           "--dlm", str(monthly[DLM_NAME]), "--inbound", str(monthly[INBOUND_NAME]),
           "--inspection", str(refs[INSPECTION_NAME]),
           "--agreements", str(refs[AGREEMENT_NAME]),
           "--db", str(tmp / "app.db"), "--out", str(tmp / "reports")]
    env = {**os.environ, "PYTHONIOENCODING": "utf-8"}
    proc = subprocess.run(cmd, cwd=REPO_ROOT, capture_output=True, text=True,
                          encoding="utf-8", env=env, timeout=300)
    assert proc.returncode == 0, f"CLI 失败:\n{proc.stderr}"
    summary = json.loads(proc.stdout)
    return SimpleNamespace(summary=summary, monthly=monthly, refs=refs,
                           out=tmp / "reports", report=tmp / "reports" / REPORT_NAME,
                           low_report=tmp / "reports" / LOW_REPORT_NAME)


def test_cli_full_run_produces_report(cli_run):
    s = cli_run.summary
    assert s["report_month"] == MONTH
    assert s["file_name"] == REPORT_NAME
    assert s["supplier_count"] == 3 and s["low200_count"] == 2
    assert s["review_count"] == 2 and s["missing_price_count"] == 0
    assert cli_run.report.exists()
    wb = openpyxl.load_workbook(cli_run.report)
    for name in ("供应商费用明细", "季度累计", "买家备注复核清单",
                 "供应商批次合格率", "数据校验"):
        assert name in wb.sheetnames, f"缺 sheet {name}"
    assert "低于200清单" in openpyxl.load_workbook(cli_run.low_report).sheetnames


def test_sample_1_pass_rate_crosscheck_with_inspection_file(cli_run):
    """抽查供应商①（甲，有验货）：报告「当月检验合格率」== 验货 xlsx 原始单元格独立重算。"""
    wb = openpyxl.load_workbook(cli_run.report)
    ws = _supplier_sheet(cli_run, JIA)
    reported = _labeled_value(ws, "当月检验合格率", 5)
    expected = _recount_pass_rate(cli_run.refs[INSPECTION_NAME], JIA, MONTH)
    assert expected is not None, "甲在 2026-07 应有验货记录"
    assert abs(reported - expected) < 1e-9
    assert abs(reported - 19 / 20) < 1e-9          # 黄金数字 19 合格 / 20 总批


def test_sample_2_sku_qty_times_price_equals_amount(cli_run):
    """抽查供应商②（全部 3 家）：每条有价 SKU 行 (E+F+G)×I=J；甲的件数与订单独立重算一致。"""
    for supplier in (JIA, YI, BING):
        for row in _sku_rows(_supplier_sheet(cli_run, supplier)):
            qty = (row["qty_def"] or 0) + (row["qty_mp"] or 0) + (row["qty_qu"] or 0)
            if row["price"] is not None:
                assert row["amount"] is not None, (supplier, row)
                assert abs(row["amount"] - qty * row["price"]) < 1e-9, (supplier, row)

    # 甲 SKU001：件数用 FBA/FBM 原始订单独立重算交叉校验（3 DEF + 1 MP，非质量单不计）
    row = _sku_rows(_supplier_sheet(cli_run, JIA))[0]
    recount = _recount_quality_qty(cli_run.monthly[FBA_NAME], cli_run.monthly[FBM_NAME],
                                   "SKU001")
    assert (row["qty_def"], row["qty_mp"], row["qty_qu"]) == (
        recount["DEFECTIVE"], recount["MISSING_PARTS"], recount["QUALITY_UNACCEPTABLE"])
    assert row["price"] == 70.0                    # 最近入库价（07-10 的 70，非 08-02 的 99）
    assert abs(row["amount"] - 4 * 70.0) < 1e-9


def test_sample_3_second_supplier_split_rows(cli_run):
    """抽查供应商③（乙/丙，交货拆分）：note=按交货比例分摊、小数件数、两行件数守恒、各自单价。"""
    wb = openpyxl.load_workbook(cli_run.report)
    yi = _sku_rows(_supplier_sheet(cli_run, YI))[0]
    bing = _sku_rows(_supplier_sheet(cli_run, BING))[0]
    assert yi["note"] == bing["note"] == "按交货比例分摊"
    assert yi["qty_qu"] == 1 and bing["qty_qu"] == 1         # 分摊往上取整（用户拍板）
    assert yi["qty_qu"] + bing["qty_qu"] == 2                # 两家各 ceil(0.5)=1（不再守恒，向上取整）
    assert yi["price"] == 30.0 and bing["price"] == 40.0      # 二供各自取各自的价
    assert abs(yi["amount"] - 30.0) < 1e-9 and abs(bing["amount"] - 40.0) < 1e-9


# ---------- 第 2 层：真实数据 E2E（marker=integration，默认 deselect） ----------

class TestRealJuly:
    pytestmark = [
        pytest.mark.integration,
        pytest.mark.skipif(not all(p.exists() for p in REAL.values()),
                           reason="真实样例文件不在本机（D:/Downloads、飞书下载、data/）"),
    ]

    @pytest.fixture(scope="class")
    def july_real_run(self, tmp_path_factory):
        """真实 7 月文件全量跑：产出与 SQLite 全部隔离在 tmp，不动项目 data/。"""
        tmp = tmp_path_factory.mktemp("real_july")
        ref = ReferenceData(loaders.load_inspection(str(REAL["inspection"])),
                            loaders.load_agreements(str(REAL["agreements"])))
        store = pipeline.Store(str(tmp / "app.db"))
        summary = pipeline.run_month("2026-07", str(REAL["fba"]), str(REAL["fbm"]),
                                     str(REAL["dlm"]), str(REAL["inbound"]),
                                     ref, str(tmp / "reports"), store=store)
        return summary, store

    def test_real_run_produces_report(self, july_real_run):
        summary, _ = july_real_run
        assert Path(summary.report_path).exists()
        assert summary.file_name == REPORT_NAME
        assert summary.supplier_count >= 1   # 采购入库单样例仅覆盖 07-31~08-04，缺价多为预期

    def test_real_report_sheets_present(self, july_real_run):
        summary, _ = july_real_run
        wb = openpyxl.load_workbook(summary.report_path)
        for name in ("供应商费用明细", "低于200清单", "季度累计", "买家备注复核清单",
                     "供应商批次合格率", "数据校验"):
            assert name in wb.sheetnames, f"缺 sheet {name}"

    def test_real_pass_rate_crosscheck_with_inspection_file(self, july_real_run):
        """交叉校验：每家有验货的供应商，报告合格率 == 验货文件 openpyxl 独立重算。"""
        _, store = july_real_run
        checked = 0
        for row in store.month_suppliers("2026-07"):
            if row["pass_rate"] is None:
                continue
            expected = _recount_pass_rate(REAL["inspection"], row["supplier"], "2026-07")
            assert expected is not None, row["supplier"]
            assert abs(row["pass_rate"] - expected) < 1e-9, row["supplier"]
            checked += 1
        assert checked >= 1

    def test_real_qty_times_price_internal_consistency(self, july_real_run):
        """抽查口径推广到全表：每个结算清单有价数据行 (E+F+G)×I≈J。"""
        summary, _ = july_real_run
        wb = openpyxl.load_workbook(summary.report_path)
        checked = 0
        for ws in wb.worksheets:
            if ws["A2"].value and str(ws["A2"].value).startswith("供应商名称："):
                for row in _sku_rows(ws):
                    if row["price"] is None:
                        continue
                    qty = (row["qty_def"] or 0) + (row["qty_mp"] or 0) + (row["qty_qu"] or 0)
                    assert row["amount"] is not None, (ws.title, row)
                    assert abs(row["amount"] - qty * row["price"]) < 0.01, (ws.title, row)
                    checked += 1
        assert checked >= 1

    def test_real_batch_matrix_matches_inspection_file(self, july_real_run):
        """矩阵 sheet 的 2026-07 列与验货文件 openpyxl 独立重算交叉校验。"""
        summary, _ = july_real_run
        wb = openpyxl.load_workbook(summary.report_path)
        ws = wb["供应商批次合格率"]
        header = [c.value for c in ws[1]]
        col = header.index("2026-07") + 1           # 该月「总批数」列
        checked = 0
        for r in range(3, ws.max_row + 1):
            sup = ws.cell(row=r, column=1).value
            total = ws.cell(row=r, column=col).value
            if not sup or not total:            # 备注行/缺月行在月列上无数字
                continue
            failed = ws.cell(row=r, column=col + 1).value
            rate = ws.cell(row=r, column=col + 2).value
            expected = _recount_pass_rate(REAL["inspection"], str(sup), "2026-07")
            assert expected is not None, sup
            assert abs(rate - (total - failed) / total) < 1e-9, sup
            assert abs(rate - expected) < 1e-9, sup
            checked += 1
        assert checked >= 1

    def test_real_review_lines_have_comments(self, july_real_run):
        summary, _ = july_real_run
        if summary.review_count:
            wb = openpyxl.load_workbook(summary.report_path)
            ws = wb["买家备注复核清单"]
            assert ws.cell(row=2, column=4).value   # 首条买家备注非空
