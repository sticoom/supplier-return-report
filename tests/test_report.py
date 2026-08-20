"""T4：engine/report.py —— short_name 简称 + write_report 复刻手工版式。"""
from pathlib import Path

import openpyxl

from engine import models as m
from engine.report import short_name, write_report


def _mk_supplier(name, deduction=280.0, agreement="V3版", pass_rate=0.9661,
                 coef=0.2, undertaken=56):
    lines = [m.SkuLine("SKU001", 100, 10, 3, 1, 0, 0.04, 70.0, 280.0, "")]
    return m.SupplierResult(name, deduction, pass_rate, agreement, coef,
                            undertaken, deduction < 200, lines)


def _mk_data(month="2026-07"):
    data = m.ReportData(month)
    data.suppliers = [_mk_supplier("东莞市甲五金制品有限公司")]
    data.low200 = [_mk_supplier("台州乙塑料制品有限公司", 15.0, "否", 1.0, 1.0, 15)]
    data.quarterly = [m.QuarterRow("2026-07", "台州乙塑料制品有限公司", 15.0, 15, False),
                      m.QuarterRow("季度小计", "台州乙塑料制品有限公司", 15.0, 15, True)]
    data.review = [m.ReviewLine("A1", "SKU001", "DEFECTIVE(存在瑕疵)", "破了",
                                 "2026-07-05", "东莞市甲五金制品有限公司")]
    data.validation = [m.ValidationItem("缺单价", "SKU009 @ 丙 无入库单价")]
    data.batch_matrix = {
        "东莞市甲五金制品有限公司": {"2026-06": (8, 0), "2026-07": (20, 1)},
        "台州乙塑料制品有限公司": {"2026-07": (10, 0)},
    }
    return data


def test_short_name_strips_region_and_suffix():
    assert short_name("东莞市甲五金制品有限公司") == "甲五金制品"
    assert short_name("台州乙塑料制品有限公司") == "乙塑料制品"
    assert short_name("某公司") == "某"


def test_write_report_file_name_and_sheet_order(tmp_path):
    path = write_report(str(tmp_path), _mk_data())
    assert Path(path).name == "2026年7月供应商质量退货金额汇总表.xlsx"
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["供应商费用明细", "甲五金制品", "乙塑料制品", "低于200清单",
                             "季度累计", "买家备注复核清单", "供应商批次合格率", "数据校验"]


def test_supplier_sheet_layout_matches_manual_template(tmp_path):
    path = write_report(str(tmp_path), _mk_data())
    wb = openpyxl.load_workbook(path)
    ws = wb["甲五金制品"]
    assert ws["A1"].value == " 2026 年 7 月供应商质量退货金额汇总表"
    assert ws["A2"].value == "供应商名称：东莞市甲五金制品有限公司"
    assert "E3:G3" in [str(r) for r in ws.merged_cells.ranges]
    assert ws["E4"].value.startswith("DEFECTIVE")
    assert ws["F4"].value.startswith("MISSING_PARTS")
    assert ws["G4"].value.startswith("QUALITY_UNACCEPTABLE")
    # 数据行
    assert ws["B5"].value == "SKU001" and ws["J5"].value == 280.0
    assert ws["J5"].number_format == "General" and ws["H5"].number_format == "0.00%"
    assert ws["H5"].value == "=SUM(G5+F5+E5)/C5"      # 质量退货率为 Excel 公式
    # 统计/系数/考核金额
    assert ws.cell(row=6, column=1).value == "统计金额："
    assert ws.cell(row=6, column=10).value == "=SUM(J5:J5)"
    assert ws.cell(row=7, column=1).value == "是否签署最新质量协议"
    assert ws.cell(row=7, column=5).value == "是"   # 与手工版式一致（版本在汇总表）
    assert ws.cell(row=7, column=10).value == 0.2
    assert ws.cell(row=8, column=1).value == "当月检验合格率："
    assert abs(ws.cell(row=8, column=5).value - 0.9661) < 1e-9
    assert ws.cell(row=8, column=10).value == "=J7*J6"
    # 备注与签字栏
    assert str(ws.cell(row=9, column=1).value).startswith("备注：数据来源于领星系统")
    assert ws.cell(row=10, column=1).value == "采购部"   # 紧接备注行（枫悦版式无空行）
    assert ws.cell(row=13, column=1).value == "总经理"
    assert ws.cell(row=13, column=2).value == "最终意见："


def test_summary_low200_quarterly_review_validation_sheets(tmp_path):
    path = write_report(str(tmp_path), _mk_data())
    wb = openpyxl.load_workbook(path)
    ws = wb["供应商费用明细"]
    assert [c.value for c in ws[1]][:8] == ["序号", "供应商", "应扣金额", "批次合格率",
                                            "是否签署质量协议(版本)", "承担比例(考核系数)",
                                            "应承担金额", "备注"]
    vals = [ws.cell(row=2, column=i).value for i in range(1, 9)]
    assert vals[1] == "东莞市甲五金制品有限公司" and vals[2] == 280.0 and vals[6] == 56
    low = wb["低于200清单"]
    assert low.cell(row=2, column=2).value == "台州乙塑料制品有限公司"
    assert low.cell(row=1, column=9).value == "所属季度"
    q = wb["季度累计"]
    assert q.cell(row=2, column=1).value == "2026-07"
    assert q.cell(row=3, column=1).value == "季度小计"
    rv = wb["买家备注复核清单"]
    assert rv.cell(row=2, column=4).value == "破了"
    vd = wb["数据校验"]
    assert vd.cell(row=2, column=1).value == "缺单价"


def test_batch_matrix_sheet_layout(tmp_path):
    """第 4 节第 7 个 sheet：供应商×月份矩阵（总批数/不合格批数/批次合格率），供应商用全名。"""
    path = write_report(str(tmp_path), _mk_data())
    wb = openpyxl.load_workbook(path)
    ws = wb["供应商批次合格率"]
    # 两行表头：A1:A2 供应商；每个月份行 1 合并 3 列，行 2 为三子列
    assert ws["A1"].value == "供应商"
    assert "A1:A2" in [str(r) for r in ws.merged_cells.ranges]
    assert ws["B1"].value == "2026-06" and ws["E1"].value == "2026-07"   # 月份升序
    assert "B1:D1" in [str(r) for r in ws.merged_cells.ranges]
    assert "E1:G1" in [str(r) for r in ws.merged_cells.ranges]
    assert [ws.cell(row=2, column=j).value for j in (2, 3, 4)] == [
        "总批数", "不合格批数", "批次合格率"]
    # 数据行：全名、缺月留空、合格率 = (总-不合格)/总、4 位小数百分比
    assert ws["A3"].value == "东莞市甲五金制品有限公司"
    assert (ws["B3"].value, ws["C3"].value, ws["D3"].value) == (8, 0, 1.0)
    assert (ws["E3"].value, ws["F3"].value, ws["G3"].value) == (20, 1, 19 / 20)
    assert ws["D3"].number_format == "0.0000%" and ws["G3"].number_format == "0.0000%"
    assert ws["A4"].value == "台州乙塑料制品有限公司"
    assert ws["B4"].value is None                                   # 6 月无验货 → 留空
    assert (ws["E4"].value, ws["F4"].value, ws["G4"].value) == (10, 0, 1.0)
    # 来源备注
    assert any("验货原始数据" in str(ws.cell(row=r, column=1).value or "")
               for r in range(5, ws.max_row + 1))


def test_sheet_name_truncation_and_dedupe(tmp_path):
    data = _mk_data()
    dup = _mk_supplier("深圳市甲五金制品有限公司", 300.0)   # short_name 也是 甲五金制品
    data.suppliers.append(dup)
    path = write_report(str(tmp_path), data)
    wb = openpyxl.load_workbook(path)
    names = wb.sheetnames
    assert "甲五金制品" in names and "甲五金制品(2)" in names
    for n in names:
        assert len(n) <= 31
